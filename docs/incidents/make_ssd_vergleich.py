# -*- coding: utf-8 -*-
"""Erzeugt SSD-Vergleich (SATA, 1 DWPD, PLP, 3.84TB) als XLSX im Incident-Ordner.
Preise: JACOB Live-Fetch 23.07.2026 (DC600M, PM893 verifiziert); Micron/Solidigm Marktschaetzung."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\dhenke\git\eneg-k8s-infrastructure-v2\docs\incidents\2026-07-21-s3168-ssd-ersatz-vergleich.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SSD-Vergleich SATA 1DWPD"

FONT = "Arial"
title_font = Font(name=FONT, size=14, bold=True)
sub_font = Font(name=FONT, size=9, italic=True, color="555555")
warn_font = Font(name=FONT, size=9, bold=True, color="C00000")
hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
cell_font = Font(name=FONT, size=10)
bold_cell = Font(name=FONT, size=10, bold=True)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
alt_fill = PatternFill("solid", fgColor="DDEBF7")
rec_fill = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws["A1"] = "SSD-Ersatz s3168 (k8s-prod-23) - Enterprise SATA, 1 DWPD, mit PLP, 3.84 TB"
ws["A1"].font = title_font
ws.merge_cells("A1:I1")
ws["A2"] = ("Preisstand: 23.07.2026, JACOB Live (inkl. MwSt.). Enterprise-SATA-Preise aktuell stark "
            "erhoeht (NAND-Knappheit). Bedarf: 6 Stueck. Kontext: Incident-Doku "
            "2026-07-21-s3168-raid-latency-rootcause-prod.md")
ws["A2"].font = sub_font
ws.merge_cells("A2:I2")
ws["A3"] = ("WICHTIG: Preise volatil und stark gestiegen - vor Bestellung zwingend tagesaktuell + "
            "Staffelangebot fuer 6 Stk einholen. Graumarkt/China-Importe (eBay ~450 EUR) bewusst NICHT gelistet.")
ws["A3"].font = warn_font
ws.merge_cells("A3:I3")

headers = ["Modell", "Hersteller-Nr.", "Interface / Formfaktor", "Endurance",
           "TBW", "Bewertung / Einordnung", "Bezugsquelle / Preisbasis",
           "Preis/Stk (EUR, inkl. MwSt.)", "6 Stk gesamt (EUR)"]
r = 5
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=r, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = border
    cell.alignment = center

rows = [
    ("Samsung PM893", "MZ7L33T8HBLT-00A07", "SATA 6Gb/s, 2.5\"", "1 DWPD (5J)", "7008 TBW",
     "DC-Standard, sehr verbreitet, V-NAND TLC, PLP. Aktuell guenstigste Option im Fachhandel.",
     "JACOB verifiziert 23.07.: 24 Stk, 1-2 Wo Lieferzeit", 2070.34),
    ("Kingston DC600M", "SEDC600M/3840G", "SATA 6Gb/s, 2.5\"", "1 DWPD (5J)", "7008 TBW",
     "Bewaehrte DC-SATA, 3D-TLC, HW-PLP, 5J Garantie. Direkter WD-Red-Nachfolger. Beste Verfuegbarkeit.",
     "JACOB verifiziert 23.07.: 12 Stk sofort lieferbar", 2313.93),
    ("Micron 5400 PRO", "MTFDDAK3T8TGA", "SATA 6Gb/s, 2.5\"", "1.5 DWPD (5J)", "10512 TBW",
     "176-Layer, hoechste MTTF-Klasse SATA, PLP. Mehr Endurance-Reserve (1.5 DWPD) fuer DB-WAL-Last.",
     "Marktschaetzung (nicht live verifiziert) - PRUEFEN", 2200.00),
    ("Solidigm D3-S4520", "SSDSC2KB038TZ1Z", "SATA 6Gb/s, 2.5\"", "1 DWPD (5J)", "~7000 TBW",
     "Ex-Intel DC-Reihe, sehr stabile FW, PLP, bewaehrt in VMware. DE-Verfuegbarkeit geringer.",
     "Marktschaetzung (nicht live verifiziert) - PRUEFEN", 2250.00),
]

def write_row(rz, data, fill=None, recommend=False):
    for c, val in enumerate(data, start=1):
        cell = ws.cell(row=rz, column=c, value=val)
        cell.font = bold_cell if (recommend and c == 1) else cell_font
        cell.border = border
        cell.alignment = center if c in (4, 5, 8, 9) else wrap
        if c in (8, 9):
            cell.number_format = '#,##0.00 "EUR"'
        if fill:
            cell.fill = fill

rr = r + 1
for i, row in enumerate(rows):
    total_formula = f"=H{rr}*6"
    data = list(row) + [total_formula]
    recommend = (row[0] == "Kingston DC600M")
    fill = rec_fill if recommend else (alt_fill if i % 2 == 1 else None)
    write_row(rr, data, fill=fill, recommend=recommend)
    rr += 1

note_row = rr + 1
ws.cell(row=note_row, column=1,
        value="Empfehlung: Kingston DC600M (gruen) - bewaehrtester Ersatz der WD Red SA500, sofort "
              "lieferbar. Samsung PM893 ist aktuell ~240 EUR/Stk guenstiger bei gleicher Klasse. "
              "Alle vier loesen das PLP-Grundproblem (fsync-Latenz). Micron 5400 PRO bietet mit 1.5 DWPD "
              "etwas mehr Endurance-Reserve.")
ws.cell(row=note_row, column=1).font = sub_font
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
ws.cell(row=note_row, column=1).alignment = wrap

price_note = note_row + 1
ws.cell(row=price_note, column=1,
        value="Preis-Hinweis: DC600M und PM893 sind am 23.07.2026 live bei JACOB abgefragt. Micron 5400 PRO "
              "und Solidigm D3-S4520 sind Marktschaetzungen auf gleichem Niveau und vor Kauf zu verifizieren.")
ws.cell(row=price_note, column=1).font = warn_font
ws.merge_cells(start_row=price_note, start_column=1, end_row=price_note, end_column=9)
ws.cell(row=price_note, column=1).alignment = wrap

raid_row = price_note + 1
ws.cell(row=raid_row, column=1,
        value="Hinweis RAID10 (VD238): 6 SSDs erlauben RAID10 ueber 6 Platten (3 Spiegelpaare, ~11.5 TB netto) "
              "oder 4x RAID10 + 2 Reserve/DHS. Aktuell sind es 4 SSDs. Slots am H755 sind SAS-faehig, SATA gewaehlt "
              "(identisch zum Ist-Zustand, keine Backplane-Risiken).")
ws.cell(row=raid_row, column=1).font = sub_font
ws.merge_cells(start_row=raid_row, start_column=1, end_row=raid_row, end_column=9)
ws.cell(row=raid_row, column=1).alignment = wrap

widths = [20, 22, 20, 14, 12, 42, 38, 16, 16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for rz in range(r+1, r+1+len(rows)):
    ws.row_dimensions[rz].height = 58
ws.row_dimensions[r].height = 30
ws.freeze_panes = "A6"

wb.save(OUT)
print("SAVED:", OUT)
